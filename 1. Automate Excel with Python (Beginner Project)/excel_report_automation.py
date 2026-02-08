import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

INPUT_FILE = "./1. Automate Excel with Python (Beginner Project)/sales_raw.xlsx"
OUTPUT_FILE = "./1. Automate Excel with Python (Beginner Project)/Final_Sales_Report.xlsx"

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

def style_header(ws, header_row=1):
    fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def add_borders(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = border

def main():
    # 1) Read
    df = pd.read_excel(INPUT_FILE)

    # 2) Cleaning
    df.columns = [c.strip().title() for c in df.columns]  # clean column names
    df = df.dropna(how="all")  # remove fully empty rows

    # Ensure required cols
    required = {"Date", "Product", "Region", "Sales"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in input: {missing}")

    # Convert date
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # Sales cleaning: remove commas, convert to numeric
    df["Sales"] = (
        df["Sales"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .replace("None", None)
    )
    df["Sales"] = pd.to_numeric(df["Sales"], errors="coerce").fillna(0)

    # Standardize text
    df["Product"] = df["Product"].astype(str).str.strip().str.title()
    df["Region"] = df["Region"].astype(str).str.strip().str.title()

    # Remove rows where date/product/region missing (optional; beginner-friendly)
    df = df.dropna(subset=["Date"])
    df = df[(df["Product"] != "None") & (df["Region"] != "None")]

    # 3) Summary
    total_sales = float(df["Sales"].sum())
    region_summary = df.groupby("Region", as_index=False)["Sales"].sum().sort_values("Sales", ascending=False)
    product_summary = df.groupby("Product", as_index=False)["Sales"].sum().sort_values("Sales", ascending=False)

    top_product = product_summary.iloc[0]["Product"] if len(product_summary) else "N/A"
    top_product_sales = float(product_summary.iloc[0]["Sales"]) if len(product_summary) else 0.0

    # 4) Write to Excel (using pandas first)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Clean_Data", index=False)
        # Create a summary sheet layout
        summary_rows = [
            ["Metric", "Value"],
            ["Total Sales", total_sales],
            ["Top Product", top_product],
            ["Top Product Sales", top_product_sales],
        ]
        pd.DataFrame(summary_rows[1:], columns=summary_rows[0]).to_excel(writer, sheet_name="Summary", index=False, startrow=0)

        # Add region & product summaries below
        region_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=6)
        product_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=6 + len(region_summary) + 3)

    # 5) Formatting with openpyxl
    wb = load_workbook(OUTPUT_FILE)

    ws_data = wb["Clean_Data"]
    ws_sum = wb["Summary"]

    # Style headers
    style_header(ws_data, 1)
    style_header(ws_sum, 1)

    # Align and number formatting in Summary
    for cell in ws_sum["B"]:
        cell.number_format = "#,##0"

    # Auto-fit
    auto_fit_columns(ws_data)
    auto_fit_columns(ws_sum)

    # Borders
    add_borders(ws_data)
    add_borders(ws_sum)

    # Conditional formatting: highlight high sales in Clean_Data (Sales column)
    # Find Sales column index
    headers = [c.value for c in ws_data[1]]
    sales_col_idx = headers.index("Sales") + 1
    sales_col_letter = get_column_letter(sales_col_idx)
    sales_range = f"{sales_col_letter}2:{sales_col_letter}{ws_data.max_row}"

    ws_data.conditional_formatting.add(
        sales_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["50000"],
                   fill=PatternFill("solid", fgColor="C6EFCE"))  # light green
    )

    # 6) Add chart (Region-wise) in Summary sheet
    # Region summary starts at row 7 (since startrow=6, +1 header)
    region_start_row = 7
    region_end_row = region_start_row + len(region_summary)

    chart = BarChart()
    chart.title = "Sales by Region"
    chart.y_axis.title = "Sales"
    chart.x_axis.title = "Region"

    data_ref = Reference(ws_sum, min_col=2, min_row=region_start_row, max_row=region_end_row)
    cats_ref = Reference(ws_sum, min_col=1, min_row=region_start_row + 1, max_row=region_end_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 18

    ws_sum.add_chart(chart, "E2")

    wb.save(OUTPUT_FILE)
    print(f"✅ Done! Generated: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()