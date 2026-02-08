import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, BarChart, Reference

INPUT_FILE = "input_business_data.xlsx"
OUTPUT_FILE = "Business_Report.xlsx"

def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = border

def autofit(ws, maxw=45):
    for col in ws.columns:
        mx = 0
        col_letter = get_column_letter(col[0].column)
        for c in col:
            v = "" if c.value is None else str(c.value)
            mx = max(mx, len(v))
        ws.column_dimensions[col_letter].width = min(mx + 2, maxw)

def main():
    # 1) Read all sheets
    xls = pd.ExcelFile(INPUT_FILE)
    orders = pd.read_excel(xls, "Orders")
    products = pd.read_excel(xls, "Products")
    customers = pd.read_excel(xls, "Customers")
    returns = pd.read_excel(xls, "Returns")
    targets = pd.read_excel(xls, "Targets")

    # 2) Cleaning - Customers
    customers["region"] = customers["region"].astype(str).str.strip().str.title()
    customers["city"] = customers["city"].astype(str).str.strip().str.title()
    customers["segment"] = customers["segment"].astype(str).str.strip().str.title()
    customers["signup_date"] = pd.to_datetime(customers["signup_date"], errors="coerce")

    # 3) Cleaning - Orders
    orders["order_date"] = pd.to_datetime(orders["order_date"], errors="coerce")
    orders["unit_price"] = pd.to_numeric(orders["unit_price"], errors="coerce")
    orders["qty"] = pd.to_numeric(orders["qty"], errors="coerce")
    orders["discount_pct"] = pd.to_numeric(orders["discount_pct"], errors="coerce").fillna(0)

    # Remove invalid rows (neg qty, missing criticals)
    orders = orders.dropna(subset=["order_id", "order_date", "customer_id", "product_id"])
    orders = orders[(orders["qty"] > 0)]
    orders["unit_price"] = orders["unit_price"].fillna(orders["unit_price"].median())

    # Dedup (just in case)
    orders = orders.drop_duplicates(subset=["order_id"])

    # 4) Returns integration
    returns["return_date"] = pd.to_datetime(returns["return_date"], errors="coerce")
    returns_flag = returns[["order_id"]].drop_duplicates()
    returns_flag["is_returned"] = 1

    # 5) Merge / Model Data
    model = (orders
             .merge(products, on="product_id", how="left", validate="many_to_one")
             .merge(customers, on="customer_id", how="left", validate="many_to_one")
             .merge(returns_flag, on="order_id", how="left"))
    model["is_returned"] = model["is_returned"].fillna(0).astype(int)

    # Integrity checks (orphan ids)
    orphan_products = model["category"].isna().sum()
    orphan_customers = model["region"].isna().sum()

    # 6) Calculated columns
    model["gross_sales"] = model["qty"] * model["unit_price"]
    model["discount_amt"] = model["gross_sales"] * (model["discount_pct"]/100.0)
    model["net_sales"] = model["gross_sales"] - model["discount_amt"]
    model["tax_amt"] = model["net_sales"] * model["tax_rate"].fillna(0)
    model["cogs"] = model["qty"] * model["cost"].fillna(0)
    model["profit"] = (model["net_sales"] - model["tax_amt"]) - model["cogs"]
    model["margin_pct"] = np.where(model["net_sales"] > 0, model["profit"]/model["net_sales"], 0)

    model["month"] = model["order_date"].dt.to_period("M").astype(str)
    model["quarter"] = model["order_date"].dt.to_period("Q").astype(str)
    model["week"] = model["order_date"].dt.isocalendar().week.astype(int)

    # 7) KPIs
    total_net = float(model["net_sales"].sum())
    total_profit = float(model["profit"].sum())
    margin = float(total_profit / total_net) if total_net else 0
    return_rate = float(model["is_returned"].mean())
    aov = float(model["net_sales"].mean())
    uniq_customers = int(model["customer_id"].nunique())

    kpis = pd.DataFrame([
        ["Total Net Sales", total_net],
        ["Total Profit", total_profit],
        ["Margin %", margin],
        ["Return Rate", return_rate],
        ["AOV", aov],
        ["Unique Customers", uniq_customers],
        ["Orphan Product Rows", int(orphan_products)],
        ["Orphan Customer Rows", int(orphan_customers)]
    ], columns=["Metric", "Value"])

    # 8) Pivots
    pivot_region_month = (model
        .groupby(["month","region"], dropna=False)
        .agg(net_sales=("net_sales","sum"), profit=("profit","sum"), orders=("order_id","count"), returns=("is_returned","sum"))
        .reset_index()
        .sort_values(["month","net_sales"], ascending=[True, False])
    )
    pivot_category = (model
        .groupby(["category"], dropna=False)
        .agg(net_sales=("net_sales","sum"), profit=("profit","sum"), margin_pct=("margin_pct","mean"), orders=("order_id","count"))
        .reset_index()
        .sort_values("net_sales", ascending=False)
    )

    # 9) Target vs Actual
    actual = (model.groupby(["month","region"], dropna=False)["net_sales"].sum().reset_index())
    tgt = targets.copy()
    tgt["month"] = tgt["month"].astype(str)
    tva = actual.merge(tgt, on=["month","region"], how="left")
    tva["achievement_pct"] = np.where(tva["target_sales"]>0, tva["net_sales"]/tva["target_sales"], np.nan)

    # 10) Pareto - Top products
    prod = model.groupby("product_name")["net_sales"].sum().sort_values(ascending=False).reset_index()
    prod["cum_sales"] = prod["net_sales"].cumsum()
    prod["cum_pct"] = prod["cum_sales"] / prod["net_sales"].sum()

    # 11) RFM (customer segmentation)
    ref_date = model["order_date"].max() + pd.Timedelta(days=1)
    rfm = model.groupby("customer_id").agg(
        last_purchase=("order_date","max"),
        frequency=("order_id","count"),
        monetary=("net_sales","sum")
    ).reset_index()
    rfm["recency_days"] = (ref_date - rfm["last_purchase"]).dt.days

    # simple scoring (quintiles)
    rfm["R"] = pd.qcut(rfm["recency_days"], 5, labels=[5,4,3,2,1]).astype(int)
    rfm["F"] = pd.qcut(rfm["frequency"].rank(method="first"), 5, labels=[1,2,3,4,5]).astype(int)
    rfm["M"] = pd.qcut(rfm["monetary"].rank(method="first"), 5, labels=[1,2,3,4,5]).astype(int)
    rfm["RFM_Score"] = rfm["R"]*100 + rfm["F"]*10 + rfm["M"]

    # 12) Anomalies
    # very high order value OR very high discount
    anomalies = model[(model["net_sales"] > model["net_sales"].quantile(0.995)) | (model["discount_pct"] >= 40)]
    anomalies = anomalies[["order_id","order_date","customer_id","product_name","region","net_sales","discount_pct","payment_mode"]].sort_values("net_sales", ascending=False)

    # 13) Write to Excel
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as w:
        orders.to_excel(w, "Clean_Orders", index=False)
        model.to_excel(w, "Model_Data", index=False)
        kpis.to_excel(w, "KPIs", index=False)
        pivot_region_month.to_excel(w, "Pivot_Region_Month", index=False)
        pivot_category.to_excel(w, "Pivot_Category", index=False)
        tva.to_excel(w, "Target_vs_Actual", index=False)
        prod.head(200).to_excel(w, "Pareto_Products", index=False)
        rfm.to_excel(w, "Customer_RFM", index=False)
        anomalies.to_excel(w, "Anomalies", index=False)

    # 14) Excel Formatting + Charts
    wb = load_workbook(OUTPUT_FILE)

    for s in wb.sheetnames:
        ws = wb[s]
        style_sheet(ws)
        autofit(ws)

    # Conditional formatting: Pivot_Category margin low highlight
    ws_cat = wb["Pivot_Category"]
    # find margin_pct col
    headers = [c.value for c in ws_cat[1]]
    if "margin_pct" in headers:
        col = get_column_letter(headers.index("margin_pct")+1)
        rng = f"{col}2:{col}{ws_cat.max_row}"
        ws_cat.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["0.10"],
                       fill=PatternFill("solid", fgColor="FFC7CE"))
        )

    # Chart: month-wise net_sales (aggregate from Pivot_Region_Month)
    ws_prm = wb["Pivot_Region_Month"]
    # Build a helper area on the right for month totals
    prm_df = pd.read_excel(OUTPUT_FILE, sheet_name="Pivot_Region_Month")
    month_tot = prm_df.groupby("month")["net_sales"].sum().reset_index()
    start_col = ws_prm.max_column + 2
    ws_prm.cell(row=1, column=start_col, value="month")
    ws_prm.cell(row=1, column=start_col+1, value="net_sales")
    for i, row in enumerate(month_tot.itertuples(index=False), start=2):
        ws_prm.cell(row=i, column=start_col, value=row.month)
        ws_prm.cell(row=i, column=start_col+1, value=float(row.net_sales))

    style_sheet(ws_prm, freeze="A2")

    chart = LineChart()
    chart.title = "Monthly Net Sales"
    chart.y_axis.title = "Net Sales"
    chart.x_axis.title = "Month"

    data = Reference(ws_prm, min_col=start_col+1, min_row=1, max_row=1+len(month_tot))
    cats = Reference(ws_prm, min_col=start_col, min_row=2, max_row=1+len(month_tot))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws_prm.add_chart(chart, f"{get_column_letter(start_col)}{2 + len(month_tot) + 2}")

    wb.save(OUTPUT_FILE)
    print(f"✅ Done! Generated: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()